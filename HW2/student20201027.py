#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

row_id = 1;
sum_list = []

for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v

		sum_list.append(sum_v)
	row_id += 1

#grade

sorted_list = sorted(sum_list)
sorted_list.reverse()

#학점 비율 설정
  
row_id2 = 1

cal_grade = []
cal_grade.append(round(len(sum_list) * 0.15))
cal_grade.append(round(len(sum_list) * 0.3))
cal_grade.append(round(len(sum_list) * 0.5))
cal_grade.append(round(len(sum_list) * 0.7))
cal_grade.append(round(len(sum_list) * 0.85))
print(cal_grade)
#학점 계산
for row in ws:
	if row_id2 != 1:
		if ws.cell(row = row_id2, column = 7).value > (sorted_list[cal_grade[0]]):
			ws.cell(row = row_id2, column = 8).value = 'A+'
		elif ws.cell(row = row_id2, column = 7).value > (sorted_list[cal_grade[1]]):
			ws.cell(row = row_id2, column = 8).value = 'A0'
		elif ws.cell(row = row_id2, column = 7).value > (sorted_list[cal_grade[2]]):
			ws.cell(row = row_id2, column = 8).value = 'B+'
		elif ws.cell(row = row_id2, column = 7).value > (sorted_list[cal_grade[3]]):
			ws.cell(row = row_id2, column = 8).value = 'B0'
		elif ws.cell(row = row_id2, column = 7).value > (sorted_list[cal_grade[4]]):
			ws.cell(row = row_id2, column = 8).value = 'C+'
		else:
			ws.cell(row = row_id2, column = 8).value = 'C0'
	row_id2 += 1



wb.save("student.xlsx")
